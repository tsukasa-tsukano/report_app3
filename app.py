from flask import Flask, render_template, session, request, redirect, url_for
import os
from werkzeug.utils import secure_filename
from flask import send_from_directory
import pandas as pd
import openpyxl
from openpyxl.styles.alignment import Alignment

#インスタンスの作成
app = Flask(__name__)

#暗号鍵の作成
key = os.urandom(21)
app.secret_key = key

#メイン
@app.route("/")
def index():
    #表示させたいdataを渡す
    return render_template('index.html')

@app.route('/edit', methods=["POST"])
def edit():
    #取り出す
    report = request.files['report']
    this_year = request.files["this_year"]
    last_year = request.files["last_year"]
    single_item = request.files["single_item"]
    month = request.form['month']
    souninji = request.form['souninji']
    hibetsu = request.files['hibetsu']
    day = request.form['day']
    kongetudo_uriage = request.form['kongetudo_uriage']
    kongetudo_rieki = request.form['kongetudo_rieki']

    #今年の売上合計照会のCSVファイルを取得
    this_year_csv = pd.read_csv(this_year,encoding="shift-jis")

    #客数と点数と実粗利の取得
    uriage = this_year_csv.loc[0]["税抜売上(純額)"]
    kyakusuu = this_year_csv.loc[0]["客数"]
    tensuu = this_year_csv.loc[0]["点数"]
    zituarari = this_year_csv.loc[0]["実粗利"]

    #ブックを取得
    book = openpyxl.load_workbook(report)

    #シートを取得
    sheet = book.worksheets[1]

    #シートに書き込む
    sheet["B7"] = uriage
    sheet["H6"] = kyakusuu
    sheet["J10"] = tensuu
    sheet["D7"] = zituarari

    #CSVファイルを取得（昨年）
    last_year_csv = pd.read_csv(last_year, encoding="shift-jis")

    #客数と点数と実粗利の取得
    sakunen_uriage = last_year_csv.loc[0]["税抜売上(純額)"]
    sakunen_kyakusuu = last_year_csv.loc[0]["客数"]
    sakunen_tensuu = last_year_csv.loc[0]["点数"]
    sakunen_zituarari = last_year_csv.loc[0]["実粗利"]

    #シートに書き込む
    sheet["B10"] = sakunen_uriage
    sheet["H8"] = sakunen_kyakusuu
    sheet["J11"] = sakunen_tensuu
    sheet["D10"] = sakunen_zituarari

    #先月度予算の書き込み
    sheet["B6"] = sheet['B29'].value
    sheet['D6'] = sheet['B30'].value

    #今月度予算の書き込み
    sheet['B29'] = kongetudo_uriage
    sheet['B30'] = kongetudo_rieki

    #CSVファイルを取得
    single_item_csv = pd.read_csv(single_item,encoding='cp932')

    #点数でソート
    tensuu_sort = single_item_csv.sort_values("点数", ascending=False)

    #CSVファイルに出力する
    tensuu_sort.to_csv("export.csv",encoding = 'cp932', index = False)

    #CSVファイルを取得
    csv_tensuu = pd.read_csv("export.csv",encoding='cp932')

    csv_tensuu_true = csv_tensuu.loc[1:10,["点数"]]

    #シートに書き込む
    num1 = list(range(19, 29))
    num2 = list(range(1, 11))

    for i, j in zip(num1, num2):
        sheet.cell(row=i, column=12).value = csv_tensuu_true.loc[j]["点数"]
    
    #品名に上書き
    csv_tensuu_true = csv_tensuu.loc[1:10,["品名"]]

    #シートに書き込む
    for i, j in zip(num1, num2):
        sheet.cell(row=i, column=11).value = csv_tensuu_true.loc[j]["品名"]

    #粗利でソート
    arari_sort = single_item_csv.sort_values("粗利", ascending=False)

    #CSVファイルに出力する
    arari_sort.to_csv("export.csv",encoding = 'cp932', index = False)

    #CSVファイルを取得
    arari_csv = pd.read_csv("export.csv",encoding='cp932')

    arari_csv_true = arari_csv.loc[1:10,["品名"]]

    #シートに書き込む
    for i, j in zip(num1, num2):
        sheet.cell(row=i, column=10).value = arari_csv_true.loc[j]["品名"]

    #売上でソート
    uriage_sort = single_item_csv.sort_values("売上", ascending=False)

    #CSVファイルに出力する
    uriage_sort.to_csv("export.csv",encoding = 'cp932', index = False)

    #CSVファイルを取得
    uriage_csv = pd.read_csv("export.csv",encoding='cp932')

    uriage_csv_true = uriage_csv.loc[1:10,["品名"]]

    #シートに書き込む
    for i, j in zip(num1, num2):
        sheet.cell(row=i, column=9).value = uriage_csv_true.loc[j]["品名"]

    #カレンダーのリクエストをリスト化
    list_day = day.split(", ")

    #今年の売上合計照会のCSVファイルを取得
    hibetsu_csv = pd.read_csv(hibetsu,encoding="shift-jis")

    #1行目の合計値を削除
    hibetsu_csv = hibetsu_csv.drop([0])
    #営業日数を書き込む
    sheet["B11"] = len(hibetsu_csv)
    #チラシ日を除外する
    hibetsu_csv = hibetsu_csv[~hibetsu_csv["日付"].isin(list_day)]
    #平日日数を書き込む
    sheet["D11"] = len(hibetsu_csv)

    #シートに書き込む
    sheet["B21"] = "{:,}".format(hibetsu_csv["税抜売上(純額)"].sum())
    sheet["C21"] = "{:,}".format(hibetsu_csv["実粗利"].sum())
    sheet["D21"] = "{:,}".format(hibetsu_csv["客数"].sum())
    sheet["E21"] = "{:,}".format(hibetsu_csv["点数"].sum())

    for row in sheet["B21:E21"]:
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    #月度入力
    month_list = list(month)
    month_list_slice = month_list[5:8]
    month_list_join = ''.join(month_list_slice)
    month_int = int(month_list_join)
    sheet["H3"] = month_int

    if month_int == 1:
        sheet["J3"] = "12月21日"
        sheet["L3"] = "1月20日"
        sheet["B11"] = 30

    elif month_int == 2:
        sheet["J3"] = "1月21日"
        sheet["L3"] = "2月20日"

    elif month_int == 3:
        sheet["J3"] = "2月21日"
        sheet["L3"] = "3月20日"

    elif month_int == 4:
        sheet["J3"] = "3月21日"
        sheet["L3"] = "4月20日"

    elif month_int == 5:
        sheet["J3"] = "4月21日"
        sheet["L3"] = "5月20日"

    elif month_int == 6:
        sheet["J3"] = "5月21日"
        sheet["L3"] = "6月20日"

    elif month_int == 7:
        sheet["J3"] = "6月21日"
        sheet["L3"] = "7月20日"

    elif month_int == 8:
        sheet["J3"] = "7月21日"
        sheet["L3"] = "8月20日"

    elif month_int == 9:
        sheet["J3"] = "8月21日"
        sheet["L3"] = "9月20日"

    elif month_int == 10:
        sheet["J3"] = "9月21日"
        sheet["L3"] = "10月20日"

    elif month_int == 11:
        sheet["J3"] = "10月21日"
        sheet["L3"] = "11月20日"

    elif month_int == 12:
        sheet["J3"] = "11月21日"
        sheet["L3"] = "12月20日"

    #総人時入力
    sheet["B12"] = souninji
    
    #パスを作成
    result_file_path = 'report_dir/' + '編集済月度報告書.xlsx'

    #シートを上書き保存する
    book.save(result_file_path)

    return send_from_directory('report_dir','編集済月度報告書.xlsx', as_attachment=True)

@app.route('/how_to_use')
def login():
    return render_template('how_to_use.html')

#アプリケーションの起動
if __name__ == '__main__':
    app.run(debug=True)